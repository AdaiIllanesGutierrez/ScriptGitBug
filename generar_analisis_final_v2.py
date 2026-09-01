#!/usr/bin/env python3
"""
generar_analisis_final_v2.py
Uso:
    python3 generar_analisis_final_v2.py \
        --mutation-csv ~/tesis/resultados/MUTATION/results_with_metrics.csv \
        --juampi-csv   ~/tesis/resultados/JUAMPI/results_with_metrics.csv \
        --output-xlsx  ~/tesis/resultados/Analisis_Final_v2.xlsx
"""

import argparse
import sys
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
from scipy.stats import wilcoxon
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14)
SUBTITLE_FONT = Font(name="Calibri", italic=True, size=10, color="555555")
WARN_FONT = Font(name="Calibri", bold=True, size=11, color="C00000")
NOTE_FONT = Font(name="Calibri", italic=True, size=10, color="806000")

# Mapeo de nombres "lindos" del reporte -> columna real extraída de los logs.
METRIC_LABELS = {
    "coverage__LINE": "Line Coverage (%)",
    "coverage__BRANCH": "Branch Coverage (%)",
    "coverage__WEAKMUTATION": "Mutation Score (%)",  # EvoSuite no reporta "MUTATION" fuerte, se usa weak mutation
}


# ── Vargha-Delaney A12 ────────────────────────────────────────────────────
def vargha_delaney_a12(group_a: np.ndarray, group_b: np.ndarray) -> float:
    m, n = len(group_a), len(group_b)
    if m == 0 or n == 0:
        return float("nan")
    ranks = pd.Series(np.concatenate([group_a, group_b])).rank()
    r1 = ranks[:m].sum()
    return float((r1 / m - (m + 1) / 2) / n)


def a12_label(a12: float) -> str:
    if pd.isna(a12):
        return "N/A"
    d = abs(a12 - 0.5)
    if d < 0.06:
        return "Insignificante"
    elif d < 0.14:
        return "Pequeño"
    elif d < 0.21:
        return "Mediano"
    return "Grande"


def a12_direction(a12: float, metric_higher_is_better: bool = True) -> str:
    if pd.isna(a12) or abs(a12 - 0.5) < 0.06:
        return "Sin diferencia"
    favors_juampi = a12 > 0.5
    if not metric_higher_is_better:
        favors_juampi = not favors_juampi
    return "Favorece a JUAMPI" if favors_juampi else "Favorece a MUTATION"


# ── Excel helpers ──────────────────────────────────────────────────────────
def autosize(ws, ncols: int, widths=None) -> None:
    for i in range(1, ncols + 1):
        w = widths[i - 1] if widths else 18
        ws.column_dimensions[get_column_letter(i)].width = w


def autosize_from_df(ws, df: pd.DataFrame, start_row: int) -> None:
    """Calcula el ancho de cada columna según el contenido real, no un valor fijo."""
    for i, col in enumerate(df.columns, start=1):
        max_len = max(
            [len(str(col))] + [len(str(v)) for v in df[col].astype(str).tolist()]
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 3, 55)


def write_header_row(ws, row: int, headers: list) -> None:
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=j, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_df_block(ws, df: pd.DataFrame, start_row: int) -> int:
    """Escribe un DataFrame a partir de start_row (con su propio header). Devuelve la fila siguiente libre."""
    write_header_row(ws, start_row, list(df.columns))
    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(df.columns, start=1):
            val = row[col]
            if pd.isna(val):
                val = ""
            ws.cell(row=i, column=j, value=val).font = BODY_FONT
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate
    autosize_from_df(ws, df, start_row)
    return start_row + len(df) + 1


# ── Carga y agregación ──────────────────────────────────────────────────────
def load_metrics(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path, dtype=str)
    numeric_cols = [c for c in df.columns if c.startswith("coverage__") or c.startswith("goals__")]
    for c in numeric_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


def per_bug_average(df: pd.DataFrame, metric_cols: list) -> pd.DataFrame:
    group_cols = ["bug_id", "class_under_test", "java_version"]
    agg = df.groupby(group_cols).agg(
        n_seeds=("seed", "nunique"),
        n_bug_revealing=("result", lambda s: (s == "BUG-REVEALING").sum()),
        **{c: (c, "mean") for c in metric_cols},
    ).reset_index()
    return agg


def mayoria_detecta(n_bug_revealing: int, n_seeds: int) -> str:
    """Criterio de mayoría: más de la mitad de los seeds deben dar BUG-REVEALING.
    Con 3 seeds, esto significa 2 o más. Generalizado por si algún bug tiene
    menos de 3 seeds disponibles (por RUNNER ERROR en alguno)."""
    if n_seeds == 0:
        return "SIN DATOS"
    return "SI" if n_bug_revealing > n_seeds / 2 else "NO"


def check_identical_pipeline_bug(paired: pd.DataFrame, metric_cols: list) -> bool:
    if not metric_cols or paired.empty:
        return False
    mask = pd.Series([True] * len(paired))
    for c in metric_cols:
        a = paired[f"{c}_JUAMPI"]
        b = paired[f"{c}_MUTATION"]
        mask &= (a.round(4) == b.round(4)) | (a.isna() & b.isna())
    return bool(mask.all())


def main() -> None:
    parser = argparse.ArgumentParser(description="Análisis final MUTATION vs JUAMPI (formato con detección por mayoría).")
    parser.add_argument("--mutation-csv", required=True)
    parser.add_argument("--juampi-csv", required=True)
    parser.add_argument("--output-xlsx", required=True)
    parser.add_argument("--force", action="store_true")
    args = parser.parse_args()

    mutation_csv = Path(args.mutation_csv).expanduser()
    juampi_csv = Path(args.juampi_csv).expanduser()
    output_xlsx = Path(args.output_xlsx).expanduser()

    for p in (mutation_csv, juampi_csv):
        if not p.is_file():
            sys.exit(f"ERROR: no existe {p}")

    if output_xlsx.exists() and not args.force:
        sys.exit(
            f"ERROR: {output_xlsx} ya existe. No se sobreescribe.\n"
            f"Usá --force si realmente querés reemplazarlo, o elegí otra ruta."
        )

    df_mut = load_metrics(mutation_csv)
    df_juampi = load_metrics(juampi_csv)

    metric_cols = [c for c in METRIC_LABELS if c in df_mut.columns and c in df_juampi.columns]

    agg_mut = per_bug_average(df_mut, metric_cols)
    agg_juampi = per_bug_average(df_juampi, metric_cols)

    agg_mut["JUAMPI/MUTATION detecta?"] = agg_mut.apply(
        lambda r: mayoria_detecta(r["n_bug_revealing"], r["n_seeds"]), axis=1
    )
    agg_juampi["JUAMPI/MUTATION detecta?"] = agg_juampi.apply(
        lambda r: mayoria_detecta(r["n_bug_revealing"], r["n_seeds"]), axis=1
    )

    key = ["bug_id", "class_under_test", "java_version"]
    paired = pd.merge(agg_juampi, agg_mut, on=key, suffixes=("_JUAMPI", "_MUTATION"), how="outer")

    pipeline_bug_detected = check_identical_pipeline_bug(paired.dropna(subset=[f"{c}_JUAMPI" for c in metric_cols] + [f"{c}_MUTATION" for c in metric_cols], how="any"), metric_cols)

    # ── Hoja: Detección (agregada por bug, criterio de mayoría) ────────────
    deteccion_rows = []
    for _, row in paired.iterrows():
        juampi_si = row.get("JUAMPI/MUTATION detecta?_JUAMPI", "SIN DATOS")
        mut_si = row.get("JUAMPI/MUTATION detecta?_MUTATION", "SIN DATOS")

        if juampi_si == "SIN DATOS" or mut_si == "SIN DATOS":
            acuerdo = "Datos incompletos"
        elif juampi_si == "SI" and mut_si == "SI":
            acuerdo = "Ambos detectan"
        elif juampi_si == "NO" and mut_si == "NO":
            acuerdo = "Ninguno detecta"
        elif juampi_si == "SI" and mut_si == "NO":
            acuerdo = "Solo JUAMPI"
        else:
            acuerdo = "Solo MUTATION"

        deteccion_rows.append({
            "bug_id": row["bug_id"],
            "class_under_test": row["class_under_test"],
            "java_version": row["java_version"],
            "seeds JUAMPI (bug-revealing/total)": (
                f"{int(row['n_bug_revealing_JUAMPI'])}/{int(row['n_seeds_JUAMPI'])}"
                if pd.notna(row.get("n_seeds_JUAMPI")) else ""
            ),
            "JUAMPI detecta?": juampi_si,
            "seeds MUTATION (bug-revealing/total)": (
                f"{int(row['n_bug_revealing_MUTATION'])}/{int(row['n_seeds_MUTATION'])}"
                if pd.notna(row.get("n_seeds_MUTATION")) else ""
            ),
            "MUTATION detecta?": mut_si,
            "Acuerdo": acuerdo,
        })
    deteccion_df = pd.DataFrame(deteccion_rows)

    # ── Hoja: Resumen Métricas (Line/Branch/Mutation Score lado a lado) ────
    resumen_rows = []
    for _, row in paired.iterrows():
        r = {
            "bug_id": row["bug_id"],
            "class_under_test": row["class_under_test"],
            "n_pares": min(
                row.get("n_seeds_JUAMPI", 0) or 0, row.get("n_seeds_MUTATION", 0) or 0
            ),
        }
        for c in metric_cols:
            label = METRIC_LABELS[c]
            r[f"{label} (JUAMPI)"] = round(row.get(f"{c}_JUAMPI"), 2) if pd.notna(row.get(f"{c}_JUAMPI")) else None
            r[f"{label} (MUTATION)"] = round(row.get(f"{c}_MUTATION"), 2) if pd.notna(row.get(f"{c}_MUTATION")) else None
        resumen_rows.append(r)
    resumen_df = pd.DataFrame(resumen_rows)

    # ── Hoja: Estadísticos (Wilcoxon + A12), pareado por bug ───────────────
    stats_rows = []
    stat_metrics = [(c, METRIC_LABELS[c], True) for c in metric_cols]
    # Tasa de detección también se compara (0/1 por bug, criterio de mayoría)
    paired["deteccion_bin_JUAMPI"] = (paired["JUAMPI/MUTATION detecta?_JUAMPI"] == "SI").astype(float)
    paired["deteccion_bin_MUTATION"] = (paired["JUAMPI/MUTATION detecta?_MUTATION"] == "SI").astype(float)

    for col, label, higher_better in stat_metrics + [("deteccion_bin", "Tasa de Detección de Bugs (mayoría 2/3)", True)]:
        col_j = f"{col}_JUAMPI"
        col_m = f"{col}_MUTATION"
        if col_j not in paired.columns or col_m not in paired.columns:
            continue
        sub = paired[[col_j, col_m]].dropna()
        n_pairs = len(sub)
        if n_pairs < 2:
            continue
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", category=RuntimeWarning)
                stat, p_value = wilcoxon(sub[col_j], sub[col_m], method="auto")
        except ValueError:
            stat, p_value = float("nan"), float("nan")
        a12 = vargha_delaney_a12(sub[col_j].to_numpy(), sub[col_m].to_numpy())
        stats_rows.append({
            "Métrica": label,
            "N pares": n_pairs,
            "Media JUAMPI": round(sub[col_j].mean(), 4),
            "Media MUTATION": round(sub[col_m].mean(), 4),
            "Estadístico W": stat,
            "p-valor": p_value,
            "¿Significativo? (p<0.05)": "SI" if pd.notna(p_value) and p_value < 0.05 else "NO",
            "A12": round(a12, 4) if pd.notna(a12) else float("nan"),
            "Magnitud (A12)": a12_label(a12),
            "Dirección": a12_direction(a12, higher_better),
        })
    estadisticos_df = pd.DataFrame(stats_rows)

    # ── Vista Global (contadores) ───────────────────────────────────────────
    total_corridas_juampi = len(df_juampi)
    total_corridas_mutation = len(df_mut)
    bugs_juampi = df_juampi["bug_id"].nunique()
    bugs_mutation = df_mut["bug_id"].nunique()
    n_ambos = (deteccion_df["Acuerdo"] == "Ambos detectan").sum()
    n_ninguno = (deteccion_df["Acuerdo"] == "Ninguno detecta").sum()
    n_solo_juampi = (deteccion_df["Acuerdo"] == "Solo JUAMPI").sum()
    n_solo_mutation = (deteccion_df["Acuerdo"] == "Solo MUTATION").sum()
    n_incompletos = (deteccion_df["Acuerdo"] == "Datos incompletos").sum()

    # ── Escribir el Excel ────────────────────────────────────────────────────
    wb = Workbook()

    # -- Vista Global --
    ws1 = wb.active
    ws1.title = "Vista Global"
    ws1["A1"] = "VISTA GLOBAL — JUAMPI vs MUTATION"
    ws1["A1"].font = TITLE_FONT
    resumen_global = pd.DataFrame([
        {"Métrica": "Total de corridas (bug x seed x java)", "JUAMPI": total_corridas_juampi, "MUTATION": total_corridas_mutation},
        {"Métrica": "Bugs/clases distintas evaluadas", "JUAMPI": bugs_juampi, "MUTATION": bugs_mutation},
        {"Métrica": "Bugs detectados (criterio mayoría 2/3 seeds)", "JUAMPI": (deteccion_df["JUAMPI detecta?"] == "SI").sum(), "MUTATION": (deteccion_df["MUTATION detecta?"] == "SI").sum()},
        {"Métrica": "Ambos detectan", "JUAMPI": n_ambos, "MUTATION": n_ambos},
        {"Métrica": "Ninguno detecta", "JUAMPI": n_ninguno, "MUTATION": n_ninguno},
        {"Métrica": "Solo JUAMPI detecta", "JUAMPI": n_solo_juampi, "MUTATION": n_solo_juampi},
        {"Métrica": "Solo MUTATION detecta", "JUAMPI": n_solo_mutation, "MUTATION": n_solo_mutation},
        {"Métrica": "Bugs con datos incompletos (RUNNER ERROR/TIMEOUT en alguno)", "JUAMPI": n_incompletos, "MUTATION": n_incompletos},
    ])
    write_df_block(ws1, resumen_global, start_row=3)

    if pipeline_bug_detected:
        warn_row = 3 + len(resumen_global) + 2
        ws1.cell(row=warn_row, column=1, value=(
            "⚠ ADVERTENCIA: las métricas promedio por bug son IDÉNTICAS entre "
            "MUTATION y JUAMPI para todos los bugs con datos completos. Verificar "
            "los CSV de entrada antes de confiar en este análisis."
        )).font = WARN_FONT
        ws1.cell(row=warn_row, column=1).alignment = Alignment(wrap_text=True)
        ws1.merge_cells(start_row=warn_row, start_column=1, end_row=warn_row + 3, end_column=6)

    # -- Resumen Metricas --
    ws2 = wb.create_sheet("Resumen Metricas")
    write_df_block(ws2, resumen_df, start_row=1)

    # -- Detección --
    ws3 = wb.create_sheet("Detección")
    write_df_block(ws3, deteccion_df, start_row=1)

    # -- Estadísticos --
    ws4 = wb.create_sheet("Estadísticos")
    ws4["A1"] = "ANÁLISIS ESTADÍSTICO — Wilcoxon signed-rank + Vargha-Delaney A12"
    ws4["A1"].font = TITLE_FONT
    ws4["A2"] = "H0: no hay diferencia entre JUAMPI y MUTATION. Se rechaza si p < 0.05."
    ws4["A2"].font = SUBTITLE_FONT
    next_row = write_df_block(ws4, estadisticos_df, start_row=4)

    ws4.cell(row=next_row + 1, column=1, value="Guía de magnitud A12 (Vargha & Delaney, 2000):").font = SUBTITLE_FONT
    ws4.cell(row=next_row + 2, column=1, value="|A12 - 0.5| < 0.06: Insignificante | 0.06-0.14: Pequeño | 0.14-0.21: Mediano | ≥0.21: Grande").font = SUBTITLE_FONT

    ws4.cell(row=next_row + 4, column=1, value=(
        "NOTA: 'Redundancia (%)', 'Cantidad de Aserciones' y 'Cantidad de Tests' "
        "no están incluidas -- requieren parsear los archivos .java de tests "
        "generados por EvoSuite (no solo evosuite_fixed.log). Avisar si se "
        "quiere construir esa extracción por separado."
    )).font = NOTE_FONT
    ws4.cell(row=next_row + 4, column=1).alignment = Alignment(wrap_text=True)
    ws4.merge_cells(start_row=next_row + 4, start_column=1, end_row=next_row + 6, end_column=8)

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)

    print(f"OK: análisis final generado en {output_xlsx}")
    print(f"  Bugs con fila en Detección: {len(deteccion_df)}")
    print(f"  Métricas comparadas: {[METRIC_LABELS[c] for c in metric_cols]}")
    if pipeline_bug_detected:
        print("\n⚠️  ADVERTENCIA: valores idénticos entre algoritmos detectados. Revisar 'Vista Global'.")


if __name__ == "__main__":
    main()
