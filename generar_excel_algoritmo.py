#!/usr/bin/env python3
"""
generar_excel_algoritmo.py

Uso:
    python3 generar_excel_algoritmo.py \
        --metrics-csv ~/tesis/resultados/JUAMPI/results_with_metrics.csv \
        --output-xlsx ~/tesis/resultados/JUAMPI/JUAMPI_resultados.xlsx \
        --algoritmo   JUAMPI
"""

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14)


def autosize(ws, df: pd.DataFrame, start_row: int = 1) -> None:
    for i, col in enumerate(df.columns, start=1):
        max_len = max(
            [len(str(col))] + [len(str(v)) for v in df[col].astype(str).tolist()]
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 3, 45)


def write_df(ws, df: pd.DataFrame, start_row: int = 1) -> None:
    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=j, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(df.columns, start=1):
            val = row[col]
            if pd.isna(val):
                val = ""
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = BODY_FONT

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate
    autosize(ws, df, start_row)


def build_resumen(df: pd.DataFrame) -> pd.DataFrame:
    """
    Promedia las métricas numéricas a través de los seeds, por
    bug_id + class_under_test + java_version, y agrega el conteo de
    corridas BUG-REVEALING sobre el total de seeds.
    """
    group_cols = ["bug_id", "class_under_test", "java_version"]

    numeric_cols = [
        c
        for c in df.columns
        if c.startswith("coverage__") or c.startswith("goals__")
    ]
    for c in numeric_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    agg = df.groupby(group_cols).agg(
        n_seeds=("seed", "nunique"),
        n_bug_revealing=("result", lambda s: (s == "BUG-REVEALING").sum()),
        n_runner_error=("result", lambda s: (s == "RUNNER ERROR").sum()),
        n_timeout=("result", lambda s: (s == "TIMEOUT").sum()),
        **{f"{c}_avg": (c, "mean") for c in numeric_cols},
    ).reset_index()

    agg["bug_detection_rate"] = (agg["n_bug_revealing"] / agg["n_seeds"]).round(3)

    return agg


def build_deteccion(df: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "bug_id",
        "class_under_test",
        "seed",
        "java_version",
        "fixed_result",
        "buggy_result",
        "result",
        "conclusion",
    ]
    cols = [c for c in cols if c in df.columns]
    return df[cols].copy()


def main() -> None:
    parser = argparse.ArgumentParser(description="Genera el Excel individual de un algoritmo.")
    parser.add_argument("--metrics-csv", required=True, help="CSV de salida de extraer_metricas.py")
    parser.add_argument("--output-xlsx", required=True, help="Ruta del .xlsx a generar")
    parser.add_argument("--algoritmo", required=True, choices=["MUTATION", "JUAMPI"])
    parser.add_argument("--force", action="store_true", help="Sobreescribir si el .xlsx ya existe")
    args = parser.parse_args()

    metrics_csv = Path(args.metrics_csv).expanduser()
    output_xlsx = Path(args.output_xlsx).expanduser()

    if not metrics_csv.is_file():
        sys.exit(f"ERROR: no existe {metrics_csv}. Corré primero extraer_metricas.py.")

    if output_xlsx.exists() and not args.force:
        sys.exit(
            f"ERROR: {output_xlsx} ya existe. No se sobreescribe.\n"
            f"Usá --force si realmente querés reemplazarlo, o elegí otra ruta."
        )

    df = pd.read_csv(metrics_csv, dtype=str)

    wb = Workbook()

    # --- Vista Global ---
    ws1 = wb.active
    ws1.title = "Vista Global"
    write_df(ws1, df)

    # --- Resumen (promedio por bug, a través de seeds) ---
    resumen_df = build_resumen(df.copy())
    ws2 = wb.create_sheet("Resumen")
    write_df(ws2, resumen_df)

    # --- Detección (BUG-REVEALING detalle) ---
    deteccion_df = build_deteccion(df)
    ws3 = wb.create_sheet("Detección")
    write_df(ws3, deteccion_df)

    # Portada simple con metadatos, al principio.
    ws0 = wb.create_sheet("Info", 0)
    ws0["A1"] = f"Resultados — Algoritmo {args.algoritmo}"
    ws0["A1"].font = TITLE_FONT
    ws0["A3"] = "Fuente:"
    ws0["B3"] = str(metrics_csv)
    ws0["A4"] = "Total de corridas:"
    ws0["B4"] = len(df)
    ws0["A5"] = "Bugs distintos:"
    ws0["B5"] = df["bug_id"].nunique() if "bug_id" in df.columns else ""
    ws0["A6"] = "Seeds distintos:"
    ws0["B6"] = df["seed"].nunique() if "seed" in df.columns else ""
    for row in range(3, 7):
        ws0.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=10)
        ws0.cell(row=row, column=2).font = BODY_FONT
    ws0.column_dimensions["A"].width = 22
    ws0.column_dimensions["B"].width = 60

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)
    print(f"OK: Excel generado en {output_xlsx}")
    print(f"  Vista Global : {len(df)} filas")
    print(f"  Resumen      : {len(resumen_df)} filas (bug x java_version)")
    print(f"  Detección    : {len(deteccion_df)} filas")


if __name__ == "__main__":
    main()
